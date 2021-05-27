'''
功能描述：读取testData中excel的数据，转化成pyhton识别的数据格式
实现逻辑：
1-获得excel文件路径，打开文件
2-找到sheet页
3-读取第一行数据作为key值列表
4-求列表数据的最大行数
5-循环读取剩余行数，得到value列表，循环条件就是剩余行数的数量也就是第二行到最后一行
    5.1将key列表和value列表拼接，得到字典
    5.2将字典添加到列表中，循环结束得到列表全部数据

'''

import openpyxl
import os
from common.logs import logger

class ReadExcel():
    #定义初始化属性-文件相对路径
    def __init__(self):
        #求当前文件路径
        cur_path = os.path.dirname(__file__)
        #求父路径
        fat_path = os.path.dirname(cur_path)
        #测试数据的路径
        self.file_path = fat_path+r'/testData/testData1.xlsx'
        logger.debug(self.file_path)
        self.data_list = []

    def get_data(self):
        #1、打开文件
        re = openpyxl.load_workbook(self.file_path)
        #2、读取文件sheet页
        #2.1获取sheet列表
        sheet_list = re.sheetnames
        re_sheet = re[sheet_list[0]]
        #3、读取测试数据最大行数
        rows = re_sheet.max_row
        cols = re_sheet.max_column
        #4、读取key值列表
        key_list = [re_sheet.cell(1,i).value for i in range(1,cols+1)]
        logger.debug(key_list)
        #5、循环读取value列表
        for j in range(2,rows):
            value_list = [re_sheet.cell(j,i).value for i in range(1,cols+1)]
            dict1 = {key_list[k]:value_list[k] for k in range(len(key_list))}
            self.data_list.append(dict1)

        #6、返回数据
        return self.data_list

if __name__ == '__main__':
    readexcel = ReadExcel()
    data = readexcel.get_data()
    print(data)